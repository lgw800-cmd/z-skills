#!/usr/bin/env python3
"""
md2xlsx.py - Extract all Markdown tables from a .md file and save to Excel.

Usage:
    python md2xlsx.py input.md [output.xlsx]

Features:
    - Detects all tables in a Markdown file (supports multiple tables)
    - Handles GFM table syntax with alignment markers
    - Strips inline Markdown formatting (bold, italic, code, links, images)
    - Each table becomes a separate sheet (named Table_1, Table_2, ...)
    - Auto-fits column widths, bold header row, thin borders
    - Supports tables with or without leading/trailing pipe characters

Requirements:
    pip install openpyxl
"""

import sys
import re
import os
from pathlib import Path


def strip_markdown(text: str) -> str:
    """Strip inline Markdown formatting, keeping plain text."""
    text = text.strip()
    text = re.sub(r'!\[([^\]]*)\]\([^)]*\)', r'\1', text)
    text = re.sub(r'\[([^\]]*)\]\([^)]*\)', r'\1', text)
    text = re.sub(r'\x60{1,3}([^\x60]+?)\x60{1,3}', r'\1', text)
    text = re.sub(r'\*{3}(.+?)\*{3}', r'\1', text)
    text = re.sub(r'_{3}(.+?)_{3}', r'\1', text)
    text = re.sub(r'\*{2}(.+?)\*{2}', r'\1', text)
    text = re.sub(r'_{2}(.+?)_{2}', r'\1', text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    text = re.sub(r'_(.+?)_', r'\1', text)
    text = re.sub(r'~~(.+?)~~', r'\1', text)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()


def is_separator_row(line: str) -> bool:
    """Check if a line is a table separator."""
    stripped = line.strip()
    if not stripped:
        return False
    if stripped.startswith('|'):
        stripped = stripped[1:]
    if stripped.endswith('|'):
        stripped = stripped[:-1]
    cells = [c.strip() for c in stripped.split('|')]
    if not cells:
        return False
    sep_pattern = re.compile(r'^:?-{2,}:?$')
    return all(sep_pattern.match(c) for c in cells if c)


def parse_table_line(line: str) -> list:
    """Parse a single table line into cell values."""
    stripped = line.strip()
    if stripped.startswith('|'):
        stripped = stripped[1:]
    if stripped.endswith('|'):
        stripped = stripped[:-1]
    cells = stripped.split('|')
    return [strip_markdown(c) for c in cells]


def parse_alignments(line: str) -> list:
    """Parse alignment markers from a separator line."""
    stripped = line.strip()
    if stripped.startswith('|'):
        stripped = stripped[1:]
    if stripped.endswith('|'):
        stripped = stripped[:-1]
    cells = [c.strip() for c in stripped.split('|')]
    alignments = []
    for c in cells:
        if not c:
            alignments.append('left')
        elif c.startswith(':') and c.endswith(':'):
            alignments.append('center')
        elif c.endswith(':'):
            alignments.append('right')
        else:
            alignments.append('left')
    return alignments


def normalize_row(row: list, n_cols: int) -> list:
    if len(row) < n_cols:
        return row + [''] * (n_cols - len(row))
    return row[:n_cols]


def finalize_table(table_lines: list) -> dict:
    if not table_lines:
        return None
    non_sep_lines = []
    alignments = None
    for line in table_lines:
        if is_separator_row(line):
            alignments = parse_alignments(line)
        else:
            non_sep_lines.append(line)
    if not non_sep_lines or alignments is None:
        return None
    header = parse_table_line(non_sep_lines[0])
    rows = [parse_table_line(line) for line in non_sep_lines[1:]]
    n_cols = len(header)
    rows = [normalize_row(r, n_cols) for r in rows]
    if not rows and not header:
        return None
    return {
        'header': header,
        'rows': rows,
        'alignments': alignments or ['left'] * n_cols
    }


def extract_tables(md_path: str) -> list:
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    tables = []
    current_table_lines = []
    in_table = False
    in_code_block = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith('```'):
            in_code_block = not in_code_block
            if in_table and current_table_lines:
                table = finalize_table(current_table_lines)
                if table:
                    tables.append(table)
                current_table_lines = []
                in_table = False
            continue
        if in_code_block:
            continue
        is_table_line = bool(stripped) and '|' in stripped
        if is_table_line:
            current_table_lines.append(line)
            in_table = True
        else:
            if in_table and current_table_lines:
                table = finalize_table(current_table_lines)
                if table:
                    tables.append(table)
                current_table_lines = []
                in_table = False
    if in_table and current_table_lines:
        table = finalize_table(current_table_lines)
        if table:
            tables.append(table)
    return tables


def write_xlsx(tables: list, output_path: str, md_filename: str = ''):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    if tables:
        wb.remove(wb.active)

    for idx, table in enumerate(tables, start=1):
        sheet_name = f"Table_{idx}"
        ws = wb.create_sheet(title=sheet_name)
        header = table['header']
        rows = table['rows']
        alignments = table['alignments']
        n_cols = len(header)

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_font = Font(size=10)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for col_idx, cell_val in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col_idx, value=cell_val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, cell_val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                cell.font = data_font
                cell.border = thin_border
                if col_idx - 1 < len(alignments):
                    h_align = alignments[col_idx - 1]
                    cell.alignment = Alignment(horizontal=h_align, vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for col_idx in range(1, n_cols + 1):
            max_width = len(str(header[col_idx - 1])) if col_idx - 1 < len(header) else 10
            for row in rows:
                if col_idx - 1 < len(row):
                    cell_len = len(str(row[col_idx - 1]))
                    if cell_len > max_width:
                        max_width = cell_len
            col_width = min(max_width + 4, 60)
            col_width = max(col_width, 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.freeze_panes = 'A2'

    if not tables:
        ws = wb.active
        ws['A1'] = 'No tables found in the Markdown file.'

    wb.save(output_path)
    print(f"Saved {len(tables)} table(s) to: {output_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python md2xlsx.py <input.md> [output.xlsx]", file=sys.stderr)
        print("  If output.xlsx is omitted, uses <input_name>.xlsx", file=sys.stderr)
        sys.exit(1)

    md_path = sys.argv[1]
    if not os.path.isfile(md_path):
        print(f"ERROR: File not found: {md_path}", file=sys.stderr)
        sys.exit(1)

    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        stem = Path(md_path).stem
        output_path = f"{stem}.xlsx"

    tables = extract_tables(md_path)

    if not tables:
        print(f"WARNING: No Markdown tables found in '{md_path}'.", file=sys.stderr)
        sys.exit(0)

    print(f"Found {len(tables)} table(s) in '{md_path}'.")
    for idx, t in enumerate(tables, start=1):
        print(f"  Table {idx}: {len(t['header'])} columns, {len(t['rows'])} data rows")

    write_xlsx(tables, output_path, md_filename=os.path.basename(md_path))


if __name__ == '__main__':
    main()
